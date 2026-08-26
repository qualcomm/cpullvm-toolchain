#!/usr/bin/env python3
"""
generate_twister_report.py
──────────────────────────
Compare two Zephyr Twister JUnit XML files (previous and current CI run) and
produce a formatted .xlsx report.

Usage:
    python generate_twister_report.py \
        --prev  previous_twister.xml \
        --curr  merged_twister.xml   \
        --out   twister_comparison_report.xlsx \
        [--prev-date "Jun 14"] \
        [--curr-date "Jun 15"]

The script is also called automatically by the 'Zephyr Twister Tests' GitHub
Actions workflow; the workflow passes the correct paths and date labels.
"""

import argparse
import sys
from collections import defaultdict
from datetime import datetime

from junitparser import JUnitXml
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── constants ─────────────────────────────────────────────────────────────────
PASS, FAIL, SKIP, ERROR = "PASS", "FAIL", "SKIP", "ERROR"

C_DARK_BLUE = "1F3864"
C_MID_BLUE  = "2E75B6"
C_LIGHT_BLUE= "D6E4F0"
C_WHITE     = "FFFFFF"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT  = Font(bold=True, color="FFFFFF", size=13)
SUB_FONT    = Font(bold=True, size=10)
NORM_FONT   = Font(size=10)
CENTER_ALIGN= Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)


# ── helpers ───────────────────────────────────────────────────────────────────
def hex_fill(h):
    return PatternFill("solid", start_color=h, fgColor=h)

def all_borders():
    s = Side(style="thin")
    return Border(top=s, bottom=s, left=s, right=s)

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def status_style(status):
    if status == PASS:
        return hex_fill("C6EFCE"), Font(color="375623", size=10)
    if status in (FAIL, ERROR):
        return hex_fill("FFC7CE"), Font(color="9C0006", size=10)
    return hex_fill("FFEB9C"), Font(color="9C5700", size=10)   # SKIP / other

def change_style(change):
    if change == "Fixed":
        return hex_fill("C6EFCE"), Font(color="375623", bold=True, size=10)
    if change == "Regressed":
        return hex_fill("FFC7CE"), Font(color="9C0006", bold=True, size=10)
    if change == "Still Failing":
        return hex_fill("FFEB9C"), Font(color="9C5700", bold=True, size=10)
    return hex_fill(C_WHITE), NORM_FONT


# ── XML parsing ───────────────────────────────────────────────────────────────
def _case_status(case):
    """Return PASS / FAIL / SKIP / ERROR for a junitparser TestCase."""
    from junitparser import Failure, Error, Skipped
    results = list(case.iterchildren())
    for r in results:
        if isinstance(r, Error):    return ERROR
        if isinstance(r, Failure):  return FAIL
        if isinstance(r, Skipped):  return SKIP
    return PASS

def _case_message(case):
    from junitparser import Failure, Error
    for r in case.iterchildren():
        if isinstance(r, (Failure, Error)):
            msg = getattr(r, "message", "") or ""
            text= getattr(r, "text", "")   or ""
            combined = (msg + "\n" + text).strip()
            return combined[:300]   # cap to keep cells readable
    return ""

def parse_xml(path: str) -> dict:
    """Return {(suite_name, test_name, classname): (status, message)}"""
    xml = JUnitXml.fromfile(path)
    results = {}
    # JUnitXml may be a TestSuite or a TestSuites container
    suites = xml if hasattr(xml, "__iter__") else [xml]
    for suite in suites:
        suite_name = suite.name or ""
        for case in suite:
            key = (suite_name, case.name or "", case.classname or "")
            results[key] = (_case_status(case), _case_message(case))
    return results


# ── classification ────────────────────────────────────────────────────────────
def classify(prev, curr):
    prev_bad = prev in (FAIL, ERROR, SKIP)
    curr_bad = curr in (FAIL, ERROR, SKIP)
    if prev_bad and not curr_bad:  return "Fixed"
    if not prev_bad and curr_bad:  return "Regressed"
    if prev_bad and curr_bad:      return "Still Failing"
    return "Pass"

def build_rows(prev_data: dict, curr_data: dict) -> list:
    all_keys = set(prev_data) | set(curr_data)
    rows = []
    for key in sorted(all_keys):
        suite, test, classname = key
        prev_status, _         = prev_data.get(key, (SKIP, ""))
        curr_status, curr_msg  = curr_data.get(key, (SKIP, ""))
        rows.append({
            "suite":    suite,
            "test":     test,
            "platform": classname,
            "prev":     prev_status,
            "curr":     curr_status,
            "status":   classify(prev_status, curr_status),
            "msg":      curr_msg,
        })
    return rows


# ── sheet builders ────────────────────────────────────────────────────────────
def _title_row(ws, title, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"].value = title
    ws["A1"].font  = TITLE_FONT
    ws["A1"].fill  = hex_fill(C_DARK_BLUE)
    ws["A1"].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 22

def _header_row(ws, headers, hdr_color, row=2):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = hex_fill(hdr_color)
        c.alignment = CENTER_WRAP
        c.border = all_borders()
    ws.row_dimensions[row].height = 28

def build_full_comparison(wb, rows, prev_date, curr_date):
    ws = wb.create_sheet("Full Comparison")
    ws.freeze_panes = "A3"

    _title_row(ws, "Zephyr Twister Test Results – Full Comparison", 7)
    headers = ["Suite", "Test Case", "Platform",
               f"Status ({prev_date})", f"Status ({curr_date})",
               "Change", "Failure Message"]
    _header_row(ws, headers, C_MID_BLUE)
    set_col_widths(ws, {"A":24,"B":34,"C":22,"D":14,"E":14,"F":16,"G":52})

    for idx, r in enumerate(rows, 3):
        bg = C_LIGHT_BLUE if idx % 2 == 0 else C_WHITE
        for ci, val in enumerate(
            [r["suite"], r["test"], r["platform"], r["prev"], r["curr"], r["status"], r["msg"]], 1
        ):
            cell = ws.cell(row=idx, column=ci, value=val)
            cell.border = all_borders()
            cell.font   = NORM_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 7))
            if ci in (4, 5):
                f, fn = status_style(val)
                cell.fill = f; cell.font = fn; cell.alignment = CENTER_ALIGN
            elif ci == 6:
                f, fn = change_style(val)
                cell.fill = f; cell.font = fn; cell.alignment = CENTER_ALIGN
            else:
                cell.fill = hex_fill(bg)

    # mini summary box
    summary = [
        ("Total Tests",    len(rows)),
        ("Pass (curr)",    sum(1 for r in rows if r["curr"] == PASS)),
        ("Fail (curr)",    sum(1 for r in rows if r["curr"] in (FAIL, ERROR))),
        ("Skip (curr)",    sum(1 for r in rows if r["curr"] == SKIP)),
        ("Fixed",          sum(1 for r in rows if r["status"] == "Fixed")),
        ("Regressed",      sum(1 for r in rows if r["status"] == "Regressed")),
        ("Still Failing",  sum(1 for r in rows if r["status"] == "Still Failing")),
    ]
    ws["I1"].value = "Summary"
    ws["I1"].font  = Font(bold=True, size=11, color=C_DARK_BLUE)
    ws.merge_cells("I1:J1"); ws["I1"].alignment = CENTER_ALIGN
    ws["I1"].fill = hex_fill("BDD7EE")
    for i, (label, val) in enumerate(summary, 2):
        ws.cell(row=i, column=9, value=label).font = SUB_FONT
        c = ws.cell(row=i, column=10, value=val)
        c.alignment = CENTER_ALIGN; c.font = NORM_FONT
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 10


def build_platform_summary(wb, rows, prev_date, curr_date):
    ws = wb.create_sheet("Platform Summary")
    _title_row(ws, "Platform Summary", 8)
    headers = ["Platform",
               f"Pass ({prev_date})", f"Fail ({prev_date})", f"Total ({prev_date})",
               f"Pass ({curr_date})", f"Fail ({curr_date})", f"Total ({curr_date})",
               "Δ Pass"]
    _header_row(ws, headers, C_MID_BLUE)
    set_col_widths(ws, {"A":22,"B":13,"C":13,"D":13,"E":13,"F":13,"G":13,"H":12})

    plat = defaultdict(lambda: {"pp":0,"pf":0,"cp":0,"cf":0})
    for r in rows:
        p = r["platform"]
        plat[p]["pp"] += r["prev"] == PASS
        plat[p]["pf"] += r["prev"] in (FAIL, ERROR, SKIP)
        plat[p]["cp"] += r["curr"] == PASS
        plat[p]["cf"] += r["curr"] in (FAIL, ERROR, SKIP)

    for idx, (platform, s) in enumerate(sorted(plat.items()), 3):
        bg = C_LIGHT_BLUE if idx % 2 == 0 else C_WHITE
        delta = s["cp"] - s["pp"]
        vals = [platform, s["pp"], s["pf"], s["pp"]+s["pf"],
                           s["cp"], s["cf"], s["cp"]+s["cf"], delta]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=idx, column=ci, value=val)
            cell.fill = hex_fill(bg); cell.border = all_borders()
            cell.alignment = CENTER_ALIGN; cell.font = NORM_FONT
            if ci == 8 and isinstance(val, int):
                if val > 0:   cell.font = Font(color="375623", bold=True, size=10)
                elif val < 0: cell.font = Font(color="9C0006", bold=True, size=10)

    tr = len(plat) + 3
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, size=10)
    for ci in range(2, 9):
        col = get_column_letter(ci)
        if ci <= 7:
            ws.cell(row=tr, column=ci).value = f"=SUM({col}3:{col}{tr-1})"
        else:
            ws.cell(row=tr, column=ci).value = f"=E{tr}-B{tr}"
    for ci in range(1, 9):
        cell = ws.cell(row=tr, column=ci)
        cell.fill = hex_fill("BDD7EE"); cell.font = Font(bold=True, size=10)
        cell.border = all_borders(); cell.alignment = CENTER_ALIGN


def build_detail_sheet(wb, name, rows, title, hdr_color, row_color, prev_date, curr_date):
    ws = wb.create_sheet(name)
    _title_row(ws, title, 6)
    headers = ["Suite", "Test Case", "Platform",
               f"Status ({prev_date})", f"Status ({curr_date})", "Failure Message"]
    _header_row(ws, headers, hdr_color)
    set_col_widths(ws, {"A":24,"B":36,"C":22,"D":14,"E":14,"F":52})

    if not rows:
        ws.merge_cells("A3:F3")
        ws["A3"] = "✅  No entries in this category."
        ws["A3"].font = Font(italic=True, size=10)
        ws["A3"].alignment = CENTER_ALIGN
        return

    for idx, r in enumerate(rows, 3):
        bg = row_color if idx % 2 != 0 else C_WHITE
        for ci, val in enumerate(
            [r["suite"], r["test"], r["platform"], r["prev"], r["curr"], r["msg"]], 1
        ):
            cell = ws.cell(row=idx, column=ci, value=val)
            cell.border = all_borders(); cell.font = NORM_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(ci == 6))
            if ci in (4, 5):
                f, fn = status_style(val); cell.fill = f; cell.font = fn
                cell.alignment = CENTER_ALIGN
            else:
                cell.fill = hex_fill(bg)


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate Twister comparison XLSX report.")
    parser.add_argument("--prev",      required=True,  help="Previous run JUnit XML")
    parser.add_argument("--curr",      required=True,  help="Current run JUnit XML")
    parser.add_argument("--out",       default="twister_comparison_report.xlsx")
    parser.add_argument("--prev-date", default=None,   help="Label for previous run e.g. 'Jun 14'")
    parser.add_argument("--curr-date", default=None,   help="Label for current run  e.g. 'Jun 15'")
    args = parser.parse_args()

    # auto-generate date labels from today if not supplied
    today = datetime.utcnow()
    curr_date = args.curr_date or today.strftime("%b %d")
    prev_date = args.prev_date or "Prev Run"

    print(f"Parsing  prev: {args.prev}")
    prev_data = parse_xml(args.prev)
    print(f"Parsing  curr: {args.curr}")
    curr_data = parse_xml(args.curr)

    rows         = build_rows(prev_data, curr_data)
    regressions  = [r for r in rows if r["status"] == "Regressed"]
    fixed        = [r for r in rows if r["status"] == "Fixed"]
    still_fail   = [r for r in rows if r["status"] == "Still Failing"]

    print(f"  Total : {len(rows)}")
    print(f"  Fixed : {len(fixed)}  |  Regressed : {len(regressions)}  |  Still Failing : {len(still_fail)}")

    wb = Workbook()
    wb.remove(wb.active)

    build_full_comparison(wb, rows, prev_date, curr_date)
    build_platform_summary(wb, rows, prev_date, curr_date)
    build_detail_sheet(wb, "Regressions",   regressions,
                       f"⚠️  Regressions  ({prev_date} → {curr_date})",
                       "C00000", "FFE0E0", prev_date, curr_date)
    build_detail_sheet(wb, "Fixed",         fixed,
                       f"✅  Fixed  ({prev_date} → {curr_date})",
                       "375623", "E2EFDA", prev_date, curr_date)
    build_detail_sheet(wb, "Still Failing", still_fail,
                       f"🔴  Still Failing  ({prev_date} → {curr_date})",
                       "7B3F00", "FFF2CC", prev_date, curr_date)

    wb.save(args.out)
    print(f"Report saved → {args.out}")


if __name__ == "__main__":
    main()
